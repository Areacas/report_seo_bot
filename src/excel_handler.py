import logging

import openpyxl
import pandas as pd

from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import os
import glob


def get_data(file):
    xl = pd.read_excel(file)
    shop = xl.iloc[7, 2]
    article = xl.iloc[9, 2]

    start_row = 11
    date_columns = xl.columns[7:]
    result = []
    for i in range(start_row, len(xl)):
        try:
            keyword = xl.iloc[i, 0]
            frequency = xl.iloc[i, 5]
        except Exception as e:
            logging.exception("Keyword error:", e)
            continue

        for date_col in date_columns:
            try:
                value = xl.loc[i, date_col]
                numeric_value = pd.to_numeric(value, errors='coerce')

                if numeric_value is not None and not pd.isna(numeric_value):
                    date = xl.loc[start_row, date_col]
                    result.append([date, shop, article, keyword, frequency, numeric_value])
            except Exception as e:
                logging.exception("Value error:", e)
                continue
    return result


def get_data_all(file):
    xl = pd.read_excel(file)
    result = []
    keyword_row_indices = xl.index[xl['Unnamed: 0'] == 'Ключевое слово'].tolist()
    keyword_row_indices.append(len(xl))

    for i in range(len(keyword_row_indices) - 1):
        start_row = keyword_row_indices[i] + 1
        end_row = keyword_row_indices[i + 1] - 5
        block = xl.iloc[start_row:end_row]

        shop_index = keyword_row_indices[i] - 4
        article_index = keyword_row_indices[i] - 2
        date_row_index = keyword_row_indices[i]

        shop = xl.iloc[shop_index, 2]
        article = xl.iloc[article_index, 2]
        date_columns = xl.iloc[date_row_index, 7:].dropna().index.tolist()

        for j in range(len(block)):
            row = block.iloc[j]
            keyword = row[0]
            frequency = row[5]
            for date_col in date_columns:
                try:
                    value = row[date_col]
                    numeric_value = pd.to_numeric(value, errors='coerce')

                    if numeric_value is not None and not pd.isna(numeric_value):
                        date = xl.loc[date_row_index, date_col]
                        result.append([date, shop, article, keyword, frequency, numeric_value])
                except Exception as e:
                    logging.exception("Value error:", e)

    return result


def created_report(data, directory_path, user_id):
    columns = ['Дата', 'Магазин', 'Артикул', 'Ключевое слово', 'Частота WB', 'Позиция']

    df = pd.DataFrame(data, columns=columns)
    df['Артикул'] = pd.to_numeric(df['Артикул'], errors='coerce').fillna(0).astype(int)
    df_sorted = df.sort_values(by='Дата')
    df_sorted['Дата'] = pd.to_datetime(df_sorted['Дата'])

    output_file = f'reports/{user_id}/ready_{directory_path}/report.xlsx'
    df_sorted.to_excel(output_file, index=False, engine='openpyxl')

    book = openpyxl.load_workbook(output_file)
    sheet = book.active

    date_style = NamedStyle(name='custom_date_style', number_format='DD.MM')

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.style = date_style

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.column_letter != 'D':
                cell.alignment = Alignment(horizontal='center', vertical='center')

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    book.save(output_file)


def process_excels(directory, user_id):
    directory_path = directory

    file_pattern = os.path.join(f'reports/{user_id}/{directory_path}', "*.xls*")
    files = glob.glob(file_pattern)

    data = []
    for file in files:
        try:
            if directory_path == 'reports_group':
                data.extend(get_data_all(file))
            else:
                data.extend(get_data(file))
        except Exception as e:
            logging.exception(f"File error: {e}")
            continue
    try:
        created_report(data, directory_path, user_id)
    except Exception as e:
        logging.exception(f'файл не создан {e}')
        return f'файл не создан {e}'
    else:
        return 'файл создан'


def clear_directory(directory):
    files = glob.glob(os.path.join(directory, '*'))
    for f in files:
        try:
            os.remove(f)
            logging.info(f'Файл удален {f}')
        except OSError as e:
            logging.exception(f"Error: {f} : {e.strerror}")
